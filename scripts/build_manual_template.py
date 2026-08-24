#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
產生（或重新產生）「PYP蝦皮市場手動蒐集表.xlsx」空白範本。

平常不需要跑這支——範本已經存在 manual_data/ 底下，直接拿去填就好。
只有在範本被誤刪、或想要調整欄位/公式設計時，才需要重新執行這支腳本
重新產生一份全新的空白範本（**會覆蓋掉 manual_data/ 裡現有的檔案，
裡面已經填好的資料會被清空**，執行前請先備份）。

使用方式
--------
    python scripts/build_manual_template.py
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parent.parent
OUT_PATH = REPO_ROOT / "manual_data" / "PYP蝦皮市場手動蒐集表.xlsx"

FONT_NAME = "Arial"
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
EXAMPLE_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
NORMAL_FONT = Font(name=FONT_NAME, size=10)
BOLD_FONT = Font(name=FONT_NAME, bold=True, size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1F4E78")
SUBTITLE_FONT = Font(name=FONT_NAME, bold=True, size=11, color="1F4E78")
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def build() -> None:
    wb = Workbook()

    # ============================================================
    # Sheet 1: 填寫說明
    # ============================================================
    ws0 = wb.active
    ws0.title = "填寫說明"
    ws0.sheet_view.showGridLines = False
    ws0.column_dimensions["A"].width = 100

    lines = [
        ("PYP 蝦皮市場手動蒐集表", TITLE_FONT),
        ("", NORMAL_FONT),
        ("這份表格取代原本自動化爬蟲的角色，供你每次人工上蝦皮搜尋「PYP」相關關鍵字時，", NORMAL_FONT),
        ("把看到的商品資訊記錄下來。填完後傳回來（或自己執行匯入腳本），系統會照樣", NORMAL_FONT),
        ("產生季報/年報儀表板，跟原本自動化版本用的是同一套報表邏輯。", NORMAL_FONT),
        ("", NORMAL_FONT),
        ("這份檔案有兩個要填的工作表：", SUBTITLE_FONT),
        ("", NORMAL_FONT),
        ("① 商品資料", BOLD_FONT),
        ("每一列 = 一個蝦皮商品頁面。這是最主要、最可靠的資料，直接抄商品頁面上", NORMAL_FONT),
        ("公開顯示的數字（價格、累計已售出等）即可，不用猜、不用算。", NORMAL_FONT),
        ("黃色欄位是要你填的，綠色那一列是範例，實際填寫時整列刪掉或蓋掉即可。", NORMAL_FONT),
        ("", NORMAL_FONT),
        ("② 規格評價統計（選填，用來推估不同規格/選項的市場需求比例）", BOLD_FONT),
        ("蝦皮公開頁面不會顯示「這個規格賣了幾件」，只會顯示整個商品的總銷量。", NORMAL_FONT),
        ("這裡用一個變通方法：蝦皮的評價區通常會標示買家當初買的是哪個規格，", NORMAL_FONT),
        ("把最近一批評價（例如最新 30～50 則，或全部，看評價數量多寡）依規格", NORMAL_FONT),
        ("分類數一數則數，再依比例反推回總銷量，估算「這個規格大概賣了多少件」。", NORMAL_FONT),
        ("", NORMAL_FONT),
        ("這是一個推估值，不是蝦皮公開的精確數字，前提假設是「留評價的比例在各規格", NORMAL_FONT),
        ("之間差不多」，實際上可能有落差（例如某規格的人比較常留評價）。儀表板上", NORMAL_FONT),
        ("這部分的數字都會清楚標示「推估值」，不會跟①的精確銷量數字混在一起看。", NORMAL_FONT),
        ("", NORMAL_FONT),
        ("如果一個商品的評價則數太少（例如低於 10 則），這個推估的可信度會很低，", NORMAL_FONT),
        ("建議乾脆不要填這個商品的規格統計，只填①的商品總銷量就好。", NORMAL_FONT),
        ("", NORMAL_FONT),
        ("填完之後怎麼辦？", SUBTITLE_FONT),
        ("", NORMAL_FONT),
        ("把這個檔案傳給負責處理系統的人（或你自己執行 scripts/ 底下的兩支匯入", NORMAL_FONT),
        ("腳本），會自動轉成系統看得懂的格式，更新到儀表板網站上。建議頻率：", NORMAL_FONT),
        ("至少每季一次（配合季報），有餘力的話可以每月做一次，資料點越密集，", NORMAL_FONT),
        ("季報的準確度越高（尤其是新商品剛開始追蹤的那一季，數字容易偏低估）。", NORMAL_FONT),
    ]

    for i, (text, font) in enumerate(lines, start=2):
        cell = ws0.cell(row=i, column=1, value=text)
        cell.font = font
        cell.alignment = Alignment(wrap_text=False, vertical="center")

    # ============================================================
    # Sheet 2: 商品資料
    # ============================================================
    ws1 = wb.create_sheet("商品資料")
    headers1 = [
        "日期(YYYY-MM-DD)", "卡套類別", "卡套尺寸", "商品名稱", "賣家名稱",
        "商品連結", "價格(NT$)", "該商品總銷量", "備註",
    ]
    widths1 = [14, 16, 18, 30, 16, 30, 10, 12, 20]
    # 文字類欄位靠左對齊，其餘置中：B類別、C尺寸、D名稱、E賣家、F連結、I備註
    LEFT_ALIGN_COLS = (2, 3, 4, 5, 6, 9)
    for col, (h, w) in enumerate(zip(headers1, widths1), start=1):
        c = ws1.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        ws1.column_dimensions[get_column_letter(col)].width = w
    ws1.row_dimensions[1].height = 30
    ws1.freeze_panes = "A2"

    example_row1 = [
        "2026-08-24", "掀蓋式卡夾", "一般卡片尺寸", "PYP 卡套 經典款 質感卡夾", "小美精品配件",
        "https://shopee.tw/product/1234567/89012345", 199, 682, "範例列，請刪除或覆蓋",
    ]
    for col, val in enumerate(example_row1, start=1):
        c = ws1.cell(row=2, column=col, value=val)
        c.font = NORMAL_FONT
        c.fill = EXAMPLE_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="left" if col in LEFT_ALIGN_COLS else "center")

    for r in range(3, 202):
        for col in range(1, len(headers1) + 1):
            c = ws1.cell(row=r, column=col)
            c.fill = INPUT_FILL
            c.font = NORMAL_FONT
            c.border = BORDER
            c.alignment = Alignment(horizontal="left" if col in LEFT_ALIGN_COLS else "center")

    date_dv = DataValidation(type="date", operator="between", formula1="2020-01-01", formula2="2035-12-31",
                              showErrorMessage=True, errorTitle="日期格式錯誤", error="請輸入正確日期，例如 2026-08-24")
    ws1.add_data_validation(date_dv)
    date_dv.add("A3:A201")

    num_dv = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0",
                             showErrorMessage=True, errorTitle="數字錯誤", error="請輸入不小於 0 的整數")
    ws1.add_data_validation(num_dv)
    num_dv.add("H3:H201")  # 該商品總銷量

    # ============================================================
    # Sheet 3: 規格評價統計
    # ============================================================
    ws2 = wb.create_sheet("規格評價統計")
    headers2 = [
        "商品名稱\n(需與「商品資料」表完全一致)", "規格名稱", "抽樣評價則數",
        "該商品抽樣評價總則數", "此規格佔比", "商品累計已售出\n(自動帶入)", "推估此規格銷量\n(自動計算)",
    ]
    widths2 = [30, 16, 14, 18, 12, 16, 16]
    for col, (h, w) in enumerate(zip(headers2, widths2), start=1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.row_dimensions[1].height = 42
    ws2.freeze_panes = "A2"

    example_rows2 = [
        ["PYP 卡套 經典款 質感卡夾", "紅色", 18, "", "", "", ""],
        ["PYP 卡套 經典款 質感卡夾", "藍色", 9, "", "", "", ""],
        ["PYP 卡套 經典款 質感卡夾", "黑色", 13, "", "", "", ""],
    ]
    for r_off, row_vals in enumerate(example_rows2, start=2):
        for col, val in enumerate(row_vals, start=1):
            c = ws2.cell(row=r_off, column=col, value=val)
            c.font = NORMAL_FONT
            c.fill = EXAMPLE_FILL
            c.border = BORDER
            c.alignment = Alignment(horizontal="center" if col not in (1, 2) else "left")
        # D: 該商品抽樣評價總則數 = SUMIF 同商品名稱的 C 欄總和
        ws2.cell(row=r_off, column=4, value=f'=IF(A{r_off}="","",SUMIF($A$2:$A$301,A{r_off},$C$2:$C$301))')
        # E: 此規格佔比 = C / D
        ws2.cell(row=r_off, column=5,
                 value=f'=IF(OR(A{r_off}="",D{r_off}="",D{r_off}=0),"",C{r_off}/D{r_off})')
        ws2.cell(row=r_off, column=5).number_format = "0.0%"
        # F: 商品累計已售出 = 從「商品資料」表用 INDEX/MATCH 抓
        ws2.cell(row=r_off, column=6,
                 value=f'=IFERROR(INDEX(商品資料!$H:$H,MATCH(A{r_off},商品資料!$D:$D,0)),"")')
        # G: 推估此規格銷量 = E * F
        ws2.cell(row=r_off, column=7,
                 value=f'=IF(OR(E{r_off}="",F{r_off}=""),"",ROUND(E{r_off}*F{r_off},0))')
        for col in (4, 5, 6, 7):
            ws2.cell(row=r_off, column=col).font = NORMAL_FONT
            ws2.cell(row=r_off, column=col).fill = EXAMPLE_FILL
            ws2.cell(row=r_off, column=col).border = BORDER
            ws2.cell(row=r_off, column=col).alignment = Alignment(horizontal="center")

    for r in range(5, 301):
        for col in range(1, len(headers2) + 1):
            c = ws2.cell(row=r, column=col)
            c.border = BORDER
            c.font = NORMAL_FONT
            if col in (1, 2, 3):
                c.fill = INPUT_FILL
                c.alignment = Alignment(horizontal="center" if col == 2 else "left")
            else:
                c.alignment = Alignment(horizontal="center")
        ws2.cell(row=r, column=4, value=f'=IF(A{r}="","",SUMIF($A$2:$A$301,A{r},$C$2:$C$301))')
        ws2.cell(row=r, column=5, value=f'=IF(OR(A{r}="",D{r}="",D{r}=0),"",C{r}/D{r})')
        ws2.cell(row=r, column=5).number_format = "0.0%"
        ws2.cell(row=r, column=6, value=f'=IFERROR(INDEX(商品資料!$H:$H,MATCH(A{r},商品資料!$D:$D,0)),"")')
        ws2.cell(row=r, column=7, value=f'=IF(OR(E{r}="",F{r}=""),"",ROUND(E{r}*F{r},0))')

    num_dv2 = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0",
                              showErrorMessage=True, errorTitle="數字錯誤", error="請輸入不小於 0 的整數")
    ws2.add_data_validation(num_dv2)
    num_dv2.add("C2:C301")

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"已產生空白範本：{OUT_PATH}")
    print("提醒：這個範本裡的公式是用 openpyxl 寫入的，還沒有計算過的結果（快取值是空的）。")
    print("請先用 Excel 或 Google 試算表打開這個檔案存一次，讓公式重新計算過，")
    print("import_variant_estimates.py 才讀得到「規格評價統計」工作表算出來的佔比/推估銷量。")


if __name__ == "__main__":
    build()
