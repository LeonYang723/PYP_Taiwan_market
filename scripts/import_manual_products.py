#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
把「PYP蝦皮市場手動蒐集表.xlsx」的「商品資料」工作表匯入成
data/raw/<日期>.jsonl，格式跟自動化爬蟲（scraper/shopee_scraper.py）輸出的
完全相容，scripts/build_reports.py 可以直接讀取、照樣彙整季報/年報，
不需要改任何東西。

使用方式
--------
    python scripts/import_manual_products.py 路徑/到/PYP蝦皮市場手動蒐集表.xlsx

若同一份 Excel 裡有好幾個不同日期的資料列，會依日期分別寫成對應的
data/raw/<日期>.jsonl（同一天的資料寫進同一個檔案；同一天內同一個商品
出現兩次以最後一列為準）。

**同一個商品連結、拆成多個規格分開填寫時**（例如同一個蝦皮 listing 裡有
6 種規格，每種規格的價格不同，逐一填成 6 列、商品連結欄位都貼一樣的網址）：
腳本會自動偵測「同一個真實 itemid 在同一天出現多筆、但卡套類別或卡套尺寸
不同」的情況，這時每一列會依 (itemid, 卡套類別, 卡套尺寸) 算一個穩定的
代用追蹤 ID，避免互相覆蓋掉——但也因為蝦皮公開頁面不會提供「規格別」銷量，
這些列的「該商品總銷量」通常會是同一個數字（整個 listing 的合併總數，非
單一規格數字），季報會把這個數字視為個別商品處理，**加總會重複計算**，
使用時請留意（可以在「備註」欄位說明是共用總數，人工判讀時心裡有數即可，
系統目前不會自動排除這種重複）。跑完之後記得接著執行：

    python scripts/build_reports.py

重新產生 docs/data/*.json，儀表板才會反映新資料。

商品 ID 是怎麼來的？
--------------------
為了讓同一個商品在「不同輪手動蒐集」之間可以被系統認出是同一個、才能算出
季度銷量差值，這支腳本會盡量從「商品連結」欄位解析出蝦皮網址裡真正的
shopid/itemid（支援 `/product/<shopid>/<itemid>` 和 `<商品名>-i.<shopid>.<itemid>`
兩種常見網址格式）。如果連結格式辨識不出來（例如貼的是短網址），會退而
求其次用「商品名稱＋賣家名稱」算一個穩定的代用 ID——這代表**之後每一輪
填寫，同一個商品的名稱、賣家名稱都要打得盡量一致**，系統才認得出是同一個
商品。腳本執行完會列出哪些商品用了這種備援機制，提醒你留意。
"""
from __future__ import annotations

import hashlib
import json
import re
import sys
from collections import defaultdict
from datetime import datetime, timezone, timedelta
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
RAW_DATA_DIR = REPO_ROOT / "data" / "raw"
TAIPEI_TZ = timezone(timedelta(hours=8))

SHEET_NAME = "商品資料"
FIRST_DATA_ROW = 3  # 第1列是標題、第2列是範例，資料從第3列開始


def extract_ids(product_url: str, shop_url: str, item_name: str, shop_name: str):
    """回傳 (shopid, itemid, is_synthetic)。優先從網址解析真正的蝦皮 ID，
    解析不出來才用名稱算代用 ID（is_synthetic=True）。"""
    m = re.search(r"/product/(\d+)/(\d+)", product_url or "")
    if m:
        return int(m.group(1)), int(m.group(2)), False
    m = re.search(r"-i\.(\d+)\.(\d+)", product_url or "")
    if m:
        return int(m.group(1)), int(m.group(2)), False

    shopid = None
    m2 = re.search(r"/shop/(\d+)", shop_url or "")
    if m2:
        shopid = int(m2.group(1))

    key = f"{item_name.strip()}|{shop_name.strip()}"
    synthetic_itemid = -int(hashlib.md5(key.encode("utf-8")).hexdigest()[:10], 16)
    if shopid is None:
        shopid = -int(hashlib.md5(shop_name.strip().encode("utf-8")).hexdigest()[:8], 16)
    return shopid, synthetic_itemid, True


def to_float(v):
    try:
        return float(v) if v not in (None, "") else None
    except (TypeError, ValueError):
        return None


def to_int(v):
    try:
        return int(v) if v not in (None, "") else None
    except (TypeError, ValueError):
        return None


def main() -> int:
    if len(sys.argv) < 2:
        print("用法：python scripts/import_manual_products.py <xlsx 檔案路徑>")
        return 1
    xlsx_path = Path(sys.argv[1])
    if not xlsx_path.exists():
        print(f"找不到檔案：{xlsx_path}")
        return 1

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(f"這個活頁簿裡沒有「{SHEET_NAME}」工作表，請確認檔案是不是對的範本。")
        return 1
    ws = wb[SHEET_NAME]

    rows_by_date: dict[str, dict] = defaultdict(dict)
    synthetic_warnings = []
    now_iso = datetime.now(TAIPEI_TZ).isoformat()

    # 先蒐集這一天內每個真實 itemid 出現過幾種不同的 (category, size) 組合，
    # 用來判斷是不是「同一個 listing 拆成多規格填寫」的情況。
    variant_seen: dict[tuple[str, int], set] = defaultdict(set)
    parsed_rows = []

    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        values = [ws.cell(row=r, column=c).value for c in range(1, 10)]
        (date_val, category, size, item_name, shop_name, product_url,
         price, sold, _note) = values

        if not item_name or not str(item_name).strip():
            continue  # 空白列，略過

        if date_val is None:
            date_str = datetime.now(TAIPEI_TZ).strftime("%Y-%m-%d")
        elif isinstance(date_val, datetime):
            date_str = date_val.strftime("%Y-%m-%d")
        else:
            date_str = str(date_val).strip()

        item_name_s = str(item_name).strip()
        shop_name_s = str(shop_name).strip() if shop_name else ""

        shopid, itemid, is_synthetic = extract_ids(
            str(product_url or ""), "", item_name_s, shop_name_s
        )
        if is_synthetic:
            synthetic_warnings.append(f"第 {r} 列「{item_name_s}」")

        category_s = str(category).strip() if category else ""
        size_s = str(size).strip() if size else ""
        variant_seen[(date_str, itemid)].add((category_s, size_s))

        parsed_rows.append({
            "date_str": date_str,
            "row_num": r,
            "itemid": itemid,
            "shopid": shopid,
            "shop_name_s": shop_name_s,
            "item_name_s": item_name_s,
            "category_s": category_s,
            "size_s": size_s,
            "price": price,
            "sold": sold,
            "product_url": product_url,
        })

    multi_variant_warnings = []
    for p in parsed_rows:
        date_str, itemid = p["date_str"], p["itemid"]
        variants_for_item = variant_seen[(date_str, itemid)]
        track_id = itemid
        if len(variants_for_item) > 1:
            # 同一個真實 itemid、同一天出現多種規格 → 用 (itemid, category, size)
            # 算一個穩定的代用追蹤 ID，避免同一天內互相覆蓋。
            variant_key = f"{itemid}|{p['category_s']}|{p['size_s']}"
            track_id = int(hashlib.md5(variant_key.encode("utf-8")).hexdigest()[:12], 16)
            multi_variant_warnings.append(
                f"第 {p['row_num']} 列「{p['item_name_s']}」規格「{p['category_s']}/{p['size_s']}」"
            )

        row = {
            "date": date_str,
            "timestamp": now_iso,
            "keyword": "(手動輸入)",
            "itemid": track_id,
            "shopee_itemid": itemid,
            "shopid": p["shopid"],
            "shop_name": p["shop_name_s"] or None,
            "item_name": p["item_name_s"],
            "category": p["category_s"] or None,
            "size": p["size_s"] or None,
            "price_twd": to_float(p["price"]),
            "sold_recent": None,
            "historical_sold": to_int(p["sold"]),
            "rating_avg": None,
            "rating_count": None,
            "stock": None,
            "url": str(p["product_url"] or ""),
            "shop_url": None,
        }
        rows_by_date[date_str][track_id] = row  # 同一天同一商品(或同一規格)，用最後一列覆蓋

    if not rows_by_date:
        print("Excel 裡沒有任何有效的資料列（商品名稱是空的），沒有東西可以匯入。")
        return 1

    RAW_DATA_DIR.mkdir(parents=True, exist_ok=True)
    for date_str, items in rows_by_date.items():
        out_path = RAW_DATA_DIR / f"{date_str}.jsonl"
        merged = sorted(items.values(), key=lambda x: -(x["historical_sold"] or 0))
        with open(out_path, "w", encoding="utf-8") as f:
            for row in merged:
                f.write(json.dumps(row, ensure_ascii=False) + "\n")
        print(f"寫入 {out_path}（{len(merged)} 筆商品）")

    if synthetic_warnings:
        print("\n⚠️  以下商品沒辦法從連結解析出蝦皮真正的商品 ID，改用「商品名稱＋賣家名稱」算代用 ID：")
        for w in synthetic_warnings:
            print("   -", w)
        print("下次填這些商品時，商品名稱、賣家名稱請盡量打得跟這次一模一樣，")
        print("系統才能認得出是同一個商品、正確算出銷量變化。建議盡量把「商品連結」欄位填完整，")
        print("避免要靠這個備援機制。")

    if multi_variant_warnings:
        print("\nℹ️  以下商品連結跟別列重複、但卡套類別/尺寸不同，判斷是同一個 listing 拆成")
        print("   多規格填寫，已各自給一個穩定的代用追蹤 ID（不會互相覆蓋）：")
        for w in multi_variant_warnings:
            print("   -", w)
        print("   請注意：這些列的「該商品總銷量」如果是同一個數字，代表是整個 listing 的")
        print("   合併總數，季報加總時會重複計算，人工判讀時請留意。")

    print("\n完成！接下來執行「python scripts/build_reports.py」重新產生季報/年報。")
    return 0


if __name__ == "__main__":
    sys.exit(main())
