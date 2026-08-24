#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
把「PYP蝦皮市場手動蒐集表.xlsx」的「規格評價統計」工作表匯入成
docs/data/variant_estimates.json，給儀表板顯示「規格別市場需求推估」。

**這些數字是推估值，不是蝦皮公開的精確銷量**：做法是把某商品最近一批評價
依買家購買的規格分類、算出各規格的評價則數佔比，再乘上該商品的總銷量，
反推「這個規格大概賣了多少件」。前提假設是「留評價的比例在各規格之間
差不多」，實際上可能有落差，儀表板會清楚標示這是推估值。

使用方式
--------
    python scripts/import_variant_estimates.py 路徑/到/PYP蝦皮市場手動蒐集表.xlsx

注意：這支腳本讀的是 Excel 裡「規格評價統計」工作表算好的公式結果（佔比、
推估銷量），所以這個檔案要先在 Excel 或 Google 試算表裡打開/存過一次，
讓公式重新計算過，才能正確匯入。如果讀到的都是空值，通常就是這個原因。
"""
from __future__ import annotations

import json
import sys
from collections import defaultdict
from datetime import datetime, timezone, timedelta
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
OUT_PATH = REPO_ROOT / "docs" / "data" / "variant_estimates.json"
TAIPEI_TZ = timezone(timedelta(hours=8))

SHEET_NAME = "規格評價統計"
FIRST_DATA_ROW = 2


def main() -> int:
    if len(sys.argv) < 2:
        print("用法：python scripts/import_variant_estimates.py <xlsx 檔案路徑>")
        return 1
    xlsx_path = Path(sys.argv[1])
    if not xlsx_path.exists():
        print(f"找不到檔案：{xlsx_path}")
        return 1

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(f"這個活頁簿裡沒有「{SHEET_NAME}」工作表，請確認檔案是不是對的範本。")
        return 1
    ws = wb[SHEET_NAME]

    products: dict[str, dict] = {}
    skipped_uncalculated = 0

    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        item_name, variant, sample_reviews, total_reviews, share, total_sold, est_sold = (
            ws.cell(row=r, column=c).value for c in range(1, 8)
        )
        if not item_name or not str(item_name).strip():
            continue
        if not variant or sample_reviews in (None, ""):
            continue

        item_name_s = str(item_name).strip()
        if share is None or total_sold is None or est_sold is None:
            # 公式還沒被算過（Excel 沒存過/沒重新計算），這一列先跳過
            skipped_uncalculated += 1
            continue

        entry = products.setdefault(
            item_name_s,
            {"item_name": item_name_s, "total_sold": None, "sample_total_reviews": 0, "variants": []},
        )
        entry["total_sold"] = int(total_sold) if total_sold not in (None, "") else None
        entry["variants"].append(
            {
                "variant": str(variant).strip(),
                "sample_reviews": int(sample_reviews),
                "share": round(float(share), 4),
                "estimated_sold": int(est_sold),
            }
        )

    for entry in products.values():
        entry["sample_total_reviews"] = sum(v["sample_reviews"] for v in entry["variants"])
        entry["variants"].sort(key=lambda v: -v["estimated_sold"])

    result = sorted(products.values(), key=lambda e: -(e["total_sold"] or 0))

    if skipped_uncalculated:
        print(
            f"⚠️  有 {skipped_uncalculated} 列因為公式沒有算出結果（佔比/推估銷量是空的）被跳過，"
            "請確認這份 Excel 是不是有在 Excel/Google 試算表裡打開存過一次。"
        )

    if not result:
        print("沒有任何可用的規格統計資料，不會產生 variant_estimates.json（或保留舊檔案不變）。")
        return 0

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "generated_at": datetime.now(TAIPEI_TZ).isoformat(),
        "products": result,
    }
    OUT_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"完成！寫入 {OUT_PATH}（{len(result)} 個商品的規格推估資料）。")
    return 0


if __name__ == "__main__":
    sys.exit(main())
