#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
把「PYP蝦皮市場手動蒐集表.xlsx」的「其他品牌商品」工作表匯入成
data/raw_other_brands/<日期>.jsonl，格式跟 scripts/import_manual_products.py
輸出的 data/raw/<日期>.jsonl 幾乎一樣，只是多一個 brand 欄位，並且寫到
另一個獨立的資料夾（data/raw_other_brands/），不會跟 PYP 自己的資料混在
一起。scripts/build_reports.py 會分別彙整這兩批資料，輸出成兩組獨立的
報表 JSON（PYP 用原本的 docs/data/*.json；其他品牌用 docs/data/other_brands_*.json），
兩邊的季報/年報/月報邏輯完全共用同一套「累計銷量差值」演算法。

這份「其他品牌商品」工作表記錄的是：同一批賣場裡，跟 PYP 同類別
（卡套/卡磚/卡冊/收納盒等）、但不是 PYP 品牌的其他商品（例如 DEI DOW），
用來做市場比較——同一個賣場、同一個商品類別裡，PYP 的銷量/價格
跟其他品牌比起來如何。

使用方式
--------
    python scripts/import_other_brands.py 路徑/到/PYP蝦皮市場手動蒐集表.xlsx

跑完之後記得接著執行：

    python scripts/build_reports.py

重新產生 docs/data/other_brands_*.json，「其他品牌」頁面才會反映新資料。

商品 ID 的解析邏輯、同一 listing 拆多規格的處理方式，都跟
import_manual_products.py 完全相同，詳見該腳本的說明文字。
"""
from __future__ import annotations

import hashlib
import json
import re
import sys
from collections import defaultdict
from datetime import datetime, timezone, timedelta
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
RAW_DATA_DIR = REPO_ROOT / "data" / "raw_other_brands"
TAIPEI_TZ = timezone(timedelta(hours=8))

SHEET_NAME = "其他品牌商品"
FIRST_DATA_ROW = 3  # 第1列是標題、第2列是範例，資料從第3列開始


def extract_ids(product_url: str, item_name: str, shop_name: str):
    """回傳 (shopid, itemid, is_synthetic)。優先從網址解析真正的蝦皮 ID，
    解析不出來才用名稱算代用 ID（is_synthetic=True）。"""
    m = re.search(r"/product/(\d+)/(\d+)", product_url or "")
    if m:
        return int(m.group(1)), int(m.group(2)), False
    m = re.search(r"-i\.(\d+)\.(\d+)", product_url or "")
    if m:
        return int(m.group(1)), int(m.group(2)), False

    key = f"{item_name.strip()}|{shop_name.strip()}"
    synthetic_itemid = -int(hashlib.md5(key.encode("utf-8")).hexdigest()[:10], 16)
    shopid = -int(hashlib.md5(shop_name.strip().encode("utf-8")).hexdigest()[:8], 16)
    return shopid, synthetic_itemid, True


def to_float(v):
    try:
        return float(v) if v not in (None, "") else None
    except (TypeError, ValueError):
        return None


def to_int(v):
    try:
        return int(v) if v not in (None, "") else None
    except (TypeError, ValueError):
        return None


def main() -> int:
    if len(sys.argv) < 2:
        print("用法：python scripts/import_other_brands.py <xlsx 檔案路徑>")
        return 1
    xlsx_path = Path(sys.argv[1])
    if not xlsx_path.exists():
        print(f"找不到檔案：{xlsx_path}")
        return 1

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(f"這個活頁簿裡沒有「{SHEET_NAME}」工作表，請確認檔案是不是對的範本。")
        return 1
    ws = wb[SHEET_NAME]

    rows_by_date: dict[str, dict] = defaultdict(dict)
    synthetic_warnings = []
    now_iso = datetime.now(TAIPEI_TZ).isoformat()

    variant_seen: dict[tuple[str, int], set] = defaultdict(set)
    parsed_rows = []

    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        values = [ws.cell(row=r, column=c).value for c in range(1, 11)]
        (date_val, brand, category, size, item_name, shop_name, product_url,
         price, sold, _note) = values

        if not item_name or not str(item_name).strip():
            continue  # 空白列，略過

        if date_val is None:
            date_str = datetime.now(TAIPEI_TZ).strftime("%Y-%m-%d")
        elif isinstance(date_val, datetime):
            date_str = date_val.strftime("%Y-%m-%d")
        else:
            date_str = str(date_val).strip()

        item_name_s = str(item_name).strip()
        shop_name_s = str(shop_name).strip() if shop_name else ""
        brand_s = str(brand).strip() if brand else "(未標示品牌)"

        shopid, itemid, is_synthetic = extract_ids(
            str(product_url or ""), item_name_s, shop_name_s
        )
        if is_synthetic:
            synthetic_warnings.append(f"第 {r} 列「{item_name_s}」")

        category_s = str(category).strip() if category else ""
        size_s = str(size).strip() if size else ""
        variant_seen[(date_str, itemid)].add((category_s, size_s))

        parsed_rows.append({
            "date_str": date_str,
            "row_num": r,
            "itemid": itemid,
            "shopid": shopid,
            "shop_name_s": shop_name_s,
            "brand_s": brand_s,
            "item_name_s": item_name_s,
            "category_s": category_s,
            "size_s": size_s,
            "price": price,
            "sold": sold,
            "product_url": product_url,
        })

    multi_variant_warnings = []
    for p in parsed_rows:
        date_str, itemid = p["date_str"], p["itemid"]
        variants_for_item = variant_seen[(date_str, itemid)]
        track_id = itemid
        if len(variants_for_item) > 1:
            variant_key = f"{itemid}|{p['category_s']}|{p['size_s']}"
            track_id = int(hashlib.md5(variant_key.encode("utf-8")).hexdigest()[:12], 16)
            multi_variant_warnings.append(
                f"第 {p['row_num']} 列「{p['item_name_s']}」規格「{p['category_s']}/{p['size_s']}」"
            )

        row = {
            "date": date_str,
            "timestamp": now_iso,
            "keyword": "(手動輸入-其他品牌)",
            "itemid": track_id,
            "shopee_itemid": itemid,
            "shopid": p["shopid"],
            "shop_name": p["shop_name_s"] or None,
            "brand": p["brand_s"],
            "item_name": p["item_name_s"],
            "category": p["category_s"] or None,
            "size": p["size_s"] or None,
            "price_twd": to_float(p["price"]),
            "sold_recent": None,
            "historical_sold": to_int(p["sold"]),
            "rating_avg": None,
            "rating_count": None,
            "stock": None,
            "url": str(p["product_url"] or ""),
            "shop_url": None,
        }
        rows_by_date[date_str][track_id] = row

    if not rows_by_date:
        print("Excel 的「其他品牌商品」工作表裡沒有任何有效的資料列，沒有東西可以匯入。")
        return 1

    RAW_DATA_DIR.mkdir(parents=True, exist_ok=True)
    for date_str, items in rows_by_date.items():
        out_path = RAW_DATA_DIR / f"{date_str}.jsonl"
        merged = sorted(items.values(), key=lambda x: -(x["historical_sold"] or 0))
        with open(out_path, "w", encoding="utf-8") as f:
            for row in merged:
                f.write(json.dumps(row, ensure_ascii=False) + "\n")
        print(f"寫入 {out_path}（{len(merged)} 筆商品）")

    if synthetic_warnings:
        print("\n⚠️  以下商品沒辦法從連結解析出蝦皮真正的商品 ID，改用「商品名稱＋賣家名稱」算代用 ID：")
        for w in synthetic_warnings:
            print("   -", w)

    if multi_variant_warnings:
        print("\nℹ️  以下商品連結跟別列重複、但卡套類別/尺寸不同，判斷是同一個 listing 拆成")
        print("   多規格填寫，已各自給一個穩定的代用追蹤 ID（不會互相覆蓋）：")
        for w in multi_variant_warnings:
            print("   -", w)

    print("\n完成！接下來執行「python scripts/build_reports.py」重新產生報表。")
    return 0


if __name__ == "__main__":
    sys.exit(main())
